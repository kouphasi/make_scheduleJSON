import openpyxl
import json

obj_list = []

wb = openpyxl.load_workbook(r"use/教室希望.xlsx")

ws = wb.active

for row in range(2,ws.max_row+1):
    day = str(ws[f"A{row}"].value)
    class_room = str(ws[f"B{row}"].value)
    time = str(ws[f"C{row}"].value)
    obj_list.append({
        "DAY": day,
        "CLASS_ROOM": class_room,
        "TIME": time,
        "NOTICES": ""
    })


obj_json = json.dumps(obj_list)

print(obj_json)